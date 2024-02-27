from odoo import models, fields, _
from odoo.exceptions import UserError, ValidationError
import openpyxl
import base64
import logging
_logger = logging.getLogger(__name__)
import io
try:
    from odoo.tools.misc import xlsxwriter
except ImportError:
    import xlsxwriter


class FleetFinesBulkUpload(models.TransientModel):
    _name = 'vehicle.rental.bulk.upload'

    fines_file = fields.Binary(string="Upload excel file")
    updated_date = fields.Datetime(string='Uploaded Date')
    template_file = fields.Binary(string="Sample Template file")
    download_status = fields.Integer(string="Template download status", default=0)

    def download_file(self):
        return self.env.ref('fleet_analytic_accounting.download_fine_and_toll_template').report_action(self, data={}, config=False)

    def upload_file(self):
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(base64.b64decode(self.fines_file)), read_only=True)
            ws = wb.active
            for record in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=None, values_only=True):
                if record != ('Vehicle plate No.', 'Location', 'Date and Time', 'Description', 'Amount', 'Fine/Toll'):
                    raise ValidationError("Upload the fines and tolls in downloaded excel template")
            for record in ws.iter_rows(min_row=3, max_row=None, min_col=None, max_col=None, values_only=True):
                vehicle = self.env['fleet.vehicle'].search([('license_plate', '=', record[1])])
                search = self.env['vehicle.rental.fines'].search([('vehicle_id', '=', vehicle.id), ('time_date', '=', record[3])])
                if not search:

                    self.env['vehicle.rental.fines'].create({'vehicle_id': vehicle.id,
                                                             'analytic_account_id': vehicle.active_contract.id,
                                                             'location': record[2],
                                                             'time_date': record[3],
                                                             'description': record[4],
                                                             'amount': record[5],
                                                             'fine_or_toll': record[6],
                                                             })
        except ValidationError:
            raise UserError(_('Upload the fines and tolls in downloaded excel template'))
        except:
            raise UserError(_('Please insert a valid file'))


class CSVExportExciseXLSX(models.AbstractModel):
    _name = 'report.fleet_analytic_accounting.fine_and_toll_template'
    _inherit = 'report.report_xlsx.abstract'

    def generate_xlsx_report(self, workbook, data, objs):
        vehicle = self.env['fleet.vehicle'].search([('state_id', '!=', 7)])
        worksheet = workbook.add_worksheet('Fine and Tolls')
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)
        worksheet.set_column('F:F', 15)
        worksheet.set_column('G:G', 15)
        worksheet.set_column('H:H', 15)
        worksheet.set_column('I:I', 15)

        worksheet.write('A1', 'Vehicle Fines and Tolls')
        row, col = 1, 0
        worksheet.write(row, col, 'No.')
        worksheet.write(row, col+1, 'Vehicle plate No.')
        worksheet.write(row, col+2, 'Location')
        worksheet.write(row, col+3, 'Date and Time')
        worksheet.write(row, col+4, 'Description')
        worksheet.write(row, col+5, 'Amount')
        worksheet.write(row, col+6, 'Fine/Toll')

        n = 1
        for each in vehicle:
            row += 1
            col = 0
            worksheet.write(row, col, n)
            worksheet.write(row, col+1, each.license_plate)
            n = n+1
